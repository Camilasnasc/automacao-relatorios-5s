"""
Script: salvar_fotos_5s.py
Descrição: Lê a planilha GESTÃO 5S, baixa as fotos de antes e depois dos links do Google Drive,
           converte para JPG se necessário, salva nas pastas definidas e
           gera um relatório Excel consolidado ao final.
"""

import re
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from PIL import Image
from io import BytesIO
from datetime import datetime

# ── Desativa limite de tamanho de imagem do Pillow ────────
Image.MAX_IMAGE_PIXELS = None

# ── Configurações ─────────────────────────────────────────
ARQUIVO_XLSX        = "GESTÃO 5S 4.0 rec.xlsx"
PASTA_ANTES         = Path("Imagens 5S")
PASTA_DEPOIS        = Path("Imagens 5S realizadas")
RELATORIO_XLSX      = "Relatorio_Fotos_5S.xlsx"


# ── Funções auxiliares ────────────────────────────────────

def extrair_id_gdrive(url: str):
    match = re.search(r"/file/d/([a-zA-Z0-9_-]+)", url)
    if match:
        return match.group(1)
    match = re.search(r"id=([a-zA-Z0-9_-]+)", url)
    return match.group(1) if match else None


def baixar_imagem_gdrive(file_id: str):
    session = requests.Session()
    url_base = "https://drive.google.com/uc"
    params = {"export": "download", "id": file_id}

    try:
        resp = session.get(url_base, params=params, stream=True, timeout=30)
        resp.raise_for_status()

        content_type = resp.headers.get("Content-Type", "")
        if "text/html" in content_type:
            token = None
            for padrao in [r'confirm=([0-9A-Za-z_-]+)', r'"confirm","([^"]+)"']:
                match = re.search(padrao, resp.text)
                if match:
                    token = match.group(1)
                    break
            params["confirm"] = token or "t"

            uuid_match = re.search(r'uuid=([^&"]+)', resp.text)
            if uuid_match:
                params["uuid"] = uuid_match.group(1)

            resp = session.get(url_base, params=params, stream=True, timeout=30)
            resp.raise_for_status()
            content_type = resp.headers.get("Content-Type", "")

        if "text/html" in content_type:
            url_alt = f"https://drive.google.com/thumbnail?id={file_id}&sz=w2000"
            resp = session.get(url_alt, stream=True, timeout=30)
            resp.raise_for_status()

        conteudo = resp.content
        if len(conteudo) < 100:
            print(f"  [ERRO] Resposta muito pequena — verifique a permissão no Drive.")
            return None

        return conteudo

    except Exception as e:
        print(f"  [ERRO] Falha ao baixar: {e}")
        return None


def detectar_extensao(conteudo: bytes) -> str:
    assinaturas = {
        b"\xff\xd8\xff": ".jpg",
        b"\x89PNG":       ".png",
        b"GIF8":          ".gif",
        b"RIFF":          ".webp",
        b"BM":            ".bmp",
    }
    for sig, ext in assinaturas.items():
        if conteudo[:len(sig)] == sig:
            return ext
    return ".bin"


def converter_para_jpg(conteudo: bytes):
    try:
        img = Image.open(BytesIO(conteudo)).convert("RGB")
        saida = BytesIO()
        img.save(saida, format="JPEG", quality=92)
        return saida.getvalue()
    except Exception as e:
        print(f"  [ERRO] Não foi possível converter para JPG: {e}")
        return None


def processar_foto(id_str: str, url, pasta: Path, nome: str, area: str) -> dict:
    """Baixa, converte e salva uma foto. Retorna o registro do resultado."""
    destino = pasta / f"{id_str}.jpg"

    if destino.exists():
        print(f"  → já existe em '{pasta.name}', pulando.")
        return {"id": id_str, "nome": nome, "area": area,
                "status": "Já existia", "formato_original": "JPG",
                "data_download": "-"}

    if not url:
        return {"id": id_str, "nome": nome, "area": area,
                "status": "Sem foto", "formato_original": "-",
                "data_download": "-"}

    file_id = extrair_id_gdrive(str(url))
    if not file_id:
        print(f"  [AVISO] URL não reconhecida como Google Drive.")
        return {"id": id_str, "nome": nome, "area": area,
                "status": "Erro", "formato_original": "-",
                "data_download": "-"}

    conteudo = baixar_imagem_gdrive(file_id)
    if not conteudo:
        return {"id": id_str, "nome": nome, "area": area,
                "status": "Erro", "formato_original": "-",
                "data_download": "-"}

    extensao = detectar_extensao(conteudo)
    print(f"  Formato detectado: {extensao}")

    if extensao == ".bin":
        print(f"  [ERRO] Formato não identificado. Verifique a permissão no Drive.")
        return {"id": id_str, "nome": nome, "area": area,
                "status": "Erro", "formato_original": ".bin",
                "data_download": "-"}

    status_final = "Salvo"
    if extensao not in {".jpg", ".jpeg"}:
        print(f"  Convertendo {extensao} → .jpg ...")
        conteudo = converter_para_jpg(conteudo)
        if not conteudo:
            return {"id": id_str, "nome": nome, "area": area,
                    "status": "Erro", "formato_original": extensao,
                    "data_download": "-"}
        status_final = "Convertido para JPG"

    destino.write_bytes(conteudo)
    agora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    print(f"  Salvo: {destino}")

    return {"id": id_str, "nome": nome, "area": area,
            "status": status_final,
            "formato_original": extensao.upper().replace(".", ""),
            "data_download": agora}


# ── Geração do relatório Excel ────────────────────────────

def gerar_relatorio(reg_antes: list, reg_depois: list):
    wb = openpyxl.Workbook()

    def criar_aba(ws, titulo_aba, registros):
        ws.title = titulo_aba

        cor_azul     = "1F3864"
        cor_verde    = "C6EFCE"
        cor_amarelo  = "FFEB9C"
        cor_vermelho = "FFC7CE"
        cor_cinza    = "D9D9D9"
        cor_laranja  = "FCE4D6"

        fonte_titulo = Font(name="Arial", bold=True, size=14, color="FFFFFF")
        fonte_cabec  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        fonte_normal = Font(name="Arial", size=10)
        fonte_bold   = Font(name="Arial", bold=True, size=10)

        borda_fina = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        ws.merge_cells("A1:F1")
        ws["A1"] = f"RELATÓRIO CONSOLIDADO — {titulo_aba.upper()}"
        ws["A1"].font = fonte_titulo
        ws["A1"].fill = PatternFill("solid", fgColor=cor_azul)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        ws.merge_cells("A2:F2")
        ws["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws["A2"].font = Font(name="Arial", size=9, italic=True, color="888888")
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 16

        cabecalhos = ["ID", "Nome", "Área", "Status", "Formato Original", "Data/Hora do Download"]
        for col, texto in enumerate(cabecalhos, start=1):
            c = ws.cell(row=4, column=col, value=texto)
            c.font = fonte_cabec
            c.fill = PatternFill("solid", fgColor="2E75B6")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = borda_fina
        ws.row_dimensions[4].height = 20

        for i, reg in enumerate(registros, start=5):
            status = reg["status"]
            if status == "Salvo":
                cor_linha = cor_verde
            elif status == "Já existia":
                cor_linha = cor_cinza
            elif status == "Convertido para JPG":
                cor_linha = cor_amarelo
            elif status == "Sem foto":
                cor_linha = cor_laranja
            else:
                cor_linha = cor_vermelho

            valores = [reg["id"], reg.get("nome", ""), reg.get("area", ""),
                       status, reg.get("formato_original", "-"), reg.get("data_download", "-")]

            for col, val in enumerate(valores, start=1):
                c = ws.cell(row=i, column=col, value=val)
                c.font = fonte_normal
                c.fill = PatternFill("solid", fgColor=cor_linha)
                c.alignment = Alignment(horizontal="center" if col in (1, 4, 5, 6) else "left", vertical="center")
                c.border = borda_fina
            ws.row_dimensions[i].height = 16

        linha_ind   = len(registros) + 6
        total       = len(registros)
        total_salvo = sum(1 for r in registros if r["status"] in ("Salvo", "Convertido para JPG"))
        total_exist = sum(1 for r in registros if r["status"] == "Já existia")
        total_conv  = sum(1 for r in registros if r["status"] == "Convertido para JPG")
        total_erro  = sum(1 for r in registros if r["status"] == "Erro")
        total_sfoto = sum(1 for r in registros if r["status"] == "Sem foto")
        pct_ok      = f"{(total_salvo + total_exist) / total * 100:.1f}%" if total else "0%"

        indicadores = [
            ("Total de registros na planilha", total),
            ("Fotos baixadas nesta execução",  total_salvo),
            ("Fotos já existiam (puladas)",    total_exist),
            ("Fotos convertidas para JPG",     total_conv),
            ("Registros sem foto",             total_sfoto),
            ("Erros",                          total_erro),
            ("% de cobertura de fotos",        pct_ok),
        ]

        ws.merge_cells(f"A{linha_ind}:F{linha_ind}")
        ws[f"A{linha_ind}"] = "INDICADORES"
        ws[f"A{linha_ind}"].font = fonte_cabec
        ws[f"A{linha_ind}"].fill = PatternFill("solid", fgColor="1F3864")
        ws[f"A{linha_ind}"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[linha_ind].height = 20

        for j, (label, valor) in enumerate(indicadores, start=linha_ind + 1):
            cl = ws.cell(row=j, column=1, value=label)
            cv = ws.cell(row=j, column=2, value=valor)
            ws.merge_cells(f"C{j}:F{j}")
            cl.font = fonte_bold
            cv.font = fonte_normal
            cl.fill = cv.fill = PatternFill("solid", fgColor="EBF3FB")
            cl.border = cv.border = borda_fina
            cl.alignment = Alignment(horizontal="left", vertical="center")
            cv.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[j].height = 16

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 22
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 22

    # Aba 1 — fotos de antes
    criar_aba(wb.active, "Fotos Antes", reg_antes)

    # Aba 2 — fotos de depois
    ws2 = wb.create_sheet()
    criar_aba(ws2, "Fotos Depois", reg_depois)

    wb.save(RELATORIO_XLSX)
    print(f"\nRelatório gerado: {Path(RELATORIO_XLSX).resolve()}")


# ── Função principal ──────────────────────────────────────

def processar_planilha():
    if not Path(ARQUIVO_XLSX).exists():
        print(f"[ERRO] Arquivo '{ARQUIVO_XLSX}' não encontrado.")
        return

    PASTA_ANTES.mkdir(exist_ok=True)
    PASTA_DEPOIS.mkdir(exist_ok=True)
    print(f"Pasta antes : {PASTA_ANTES.resolve()}")
    print(f"Pasta depois: {PASTA_DEPOIS.resolve()}\n")

    wb = openpyxl.load_workbook(ARQUIVO_XLSX, keep_vba=True)
    ws = wb.active

    cabecalhos = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]

    try:
        col_id      = cabecalhos.index("índice")
        col_nome    = cabecalhos.index("nome")
        col_area    = cabecalhos.index("selecione a área que foi/será realizada o 5s")
        col_fotoAnt = cabecalhos.index("insira a foto de antes da atividade")
        col_fotoDep = cabecalhos.index("insira a foto após a atividade (caso já tenha sido realizado)")
    except ValueError as e:
        print(f"[ERRO] Coluna não encontrada: {e}")
        print("\nColunas disponíveis na planilha:")
        for i, nome in enumerate(cabecalhos):
            print(f"  {i}: {nome}")
        return

    reg_antes  = []
    reg_depois = []

    for linha in ws.iter_rows(min_row=2, values_only=True):
        id_val    = linha[col_id]
        url_ant   = linha[col_fotoAnt]
        url_dep   = linha[col_fotoDep]
        nome      = linha[col_nome] or ""
        area      = linha[col_area] or ""

        if id_val is None:
            continue

        id_str = str(int(id_val)) if isinstance(id_val, float) else str(id_val)

        # Foto de antes
        print(f"\nID {id_str} — foto ANTES:")
        reg_antes.append(processar_foto(id_str, url_ant, PASTA_ANTES, nome, area))

        # Foto de depois
        print(f"ID {id_str} — foto DEPOIS:")
        reg_depois.append(processar_foto(id_str, url_dep, PASTA_DEPOIS, nome, area))

    # Resumo
    def resumo(label, registros):
        salvo = sum(1 for r in registros if r["status"] in ("Salvo", "Convertido para JPG"))
        exist = sum(1 for r in registros if r["status"] == "Já existia")
        erro  = sum(1 for r in registros if r["status"] == "Erro")
        sfoto = sum(1 for r in registros if r["status"] == "Sem foto")
        print(f"\n{label}")
        print(f"  Novas salvas    : {salvo}")
        print(f"  Já existiam     : {exist}")
        print(f"  Sem foto        : {sfoto}")
        print(f"  Erros           : {erro}")

    print(f"\n{'='*40}")
    resumo("FOTOS ANTES (Imagens 5S):", reg_antes)
    resumo("FOTOS DEPOIS (Imagens 5S realizadas):", reg_depois)

    gerar_relatorio(reg_antes, reg_depois)


if __name__ == "__main__":
    processar_planilha()